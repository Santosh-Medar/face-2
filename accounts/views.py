import base64
import io
import json
import re

import cv2
import face_recognition
import numpy as np
from PIL import Image
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from django.contrib import messages
from django.contrib.auth import authenticate, get_user_model, login, logout
from django.contrib.auth.decorators import login_required
from django.core.files.base import ContentFile
from django.db import models, transaction
from django.db.models import Q
from django.http import JsonResponse, HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils.decorators import method_decorator
from django.views.decorators.csrf import csrf_exempt
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from attendance.models import AttendanceRecord
from attendance.utils import encode_face_from_image
from attendance.utils import get_mandatory_subjects, get_language_subject, get_english_subject

from .forms import (
    AdminUserUpdateForm,
    BulkUploadForm,
    NonTeachingStaffRegistrationForm,
    StudentRegistrationForm,
    TeachingStaffRegistrationForm,
)
from .models import Course, Section, StaffProfile, StudentEnrollment, StudentProfile, Subject, User

User = get_user_model()

# =============================================================================
# Helper functions
# =============================================================================

def get_mandatory_subjects(course, semester, section):
    """
    Return core subjects (excluding languages) for the given course, semester and section.
    """
    subjects = Subject.objects.filter(course=course, semester=semester)

    # Exclude language subjects (codes containing GEN, GHN, GKA, KAN, HIN)
    language_patterns = ['GEN', 'GHN', 'GKA', 'KAN', 'HIN']
    query = models.Q()
    for pattern in language_patterns:
        query |= models.Q(code__icontains=pattern)
    subjects = subjects.exclude(query)

    # For BSc, filter by section to get the correct combo
    if course.code == 'BSc':
        section_name = section.name
        if section_name == 'CBZ':
            patterns = ['BOTN', 'CHEM', 'ZOOL']
        elif section_name == 'PCM':
            patterns = ['CHEM', 'MATH', 'PHYS']
        elif section_name == 'PMCS':
            patterns = ['PHYS', 'MATH', 'COMP']
        else:
            patterns = []
        if patterns:
            query = models.Q()
            for p in patterns:
                query |= models.Q(code__icontains=p)
            subjects = subjects.filter(query)

    return subjects


def get_language_subject(course, semester, lang_code):
    """
    Return the language subject (Kannada or Hindi) for a given course and semester.
    lang_code: 'KAN' or 'HIN'
    """
    if lang_code == 'KAN':
        subject = Subject.objects.filter(
            course=course,
            semester=semester
        ).filter(
            models.Q(code__icontains='KAN') |
            models.Q(code__icontains='GKA') |
            models.Q(name__icontains='Kannada')
        ).first()
    else:  # 'HIN'
        subject = Subject.objects.filter(
            course=course,
            semester=semester
        ).filter(
            models.Q(code__icontains='GHN') |
            models.Q(code__icontains='HIN') |
            models.Q(name__icontains='Hindi')
        ).first()

    if not subject:
        raise Subject.DoesNotExist(
            f"Language subject not found for {course.code} sem {semester} lang {lang_code}"
        )
    return subject


def get_english_subject(course, semester):
    """Return the Basic English subject for a given course and semester."""
    subject = Subject.objects.filter(
        course=course,
        semester=semester,
        code__icontains='GEN'
    ).first()
    if not subject:
        raise Subject.DoesNotExist(
            f"Basic English subject not found for {course.code} sem {semester}"
        )
    return subject


def assign_subjects_to_enrollment(enrollment):
    """Assign mandatory, language and English subjects to an enrollment."""
    core = get_mandatory_subjects(enrollment.course, enrollment.semester, enrollment.section)
    lang1 = get_language_subject(enrollment.course, enrollment.semester, enrollment.language1)
    english = get_english_subject(enrollment.course, enrollment.semester)
    all_subjects = list(core) + [lang1, english]
    enrollment.subjects.add(*all_subjects)
    enrollment.save()


def handle_face_image(profile, image_file):
    """
    Save a face image to a profile and generate its encoding.
    Returns (success, message).
    """
    if not image_file:
        return False, "No image provided."

    try:
        # Reset file pointer in case it was read earlier
        image_file.seek(0)
        profile.face_image.save(
            f"{profile.user.username}_face.jpg",
            ContentFile(image_file.read()),
            save=False
        )
        profile.save()

        # Generate encoding from the saved file
        encoding = encode_face_from_image(profile.face_image.path)
        if encoding is not None:
            profile.set_face_encoding(encoding)
            profile.registration_complete = True
            profile.save()
            return True, "Face image processed successfully."
        else:
            return False, "No face detected in the image."
    except Exception as e:
        return False, f"Error processing face: {str(e)}"


def get_expected_headers(user_type):
    if user_type == 'student':
        return ['full_name', 'uucms_id', 'email', 'phone_number', 'course_code', 'semester', 'section_name', 'language1']
    elif user_type == 'lecturer':
        return ['full_name', 'username', 'email', 'phone_number', 'designation', 'course_code']
    elif user_type == 'staff':
        return ['full_name', 'username', 'email', 'phone_number', 'designation']
    return []


def process_bulk_row(user_type, row, results):
    from accounts.models import User, StudentProfile, Course, Section, StaffProfile, StudentEnrollment, Subject
    from attendance.utils import get_mandatory_subjects, get_language_subject, get_english_subject  # ✅ correct
    from django.db import transaction
    import logging
    logger = logging.getLogger(__name__)

    def clean_text(val):
        return str(val).strip() if val else ''

    print(f"process_bulk_row: user_type={user_type}, row={row}")

    # Student processing in process_bulk_row
    if user_type == 'student':
        try:
            if len(row) < 8:
                raise ValueError(f"Insufficient columns. Expected 8, got {len(row)}")
            full_name, uucms_id, email, phone, course_code, semester, section_name, language1 = row[:8]
            full_name = clean_text(full_name)
            uucms_id = clean_text(uucms_id)
            email = clean_text(email)
            phone = clean_text(phone)
            course_code = clean_text(course_code)
            semester = int(clean_text(semester)) if clean_text(semester) else None
            section_name = clean_text(section_name)
            language1 = clean_text(language1).upper()

            print(f"Student parsed: full_name={full_name}, uucms_id={uucms_id}, email={email}, course_code={course_code}, semester={semester}, section={section_name}, lang={language1}")

            if not all([full_name, uucms_id, email, course_code, semester, section_name, language1]):
                raise ValueError("Missing required fields")

            # Normalize language1
            if language1 in ['KAN', 'KANNADA']:
                language1 = 'KAN'
            elif language1 in ['HIN', 'HINDI']:
                language1 = 'HIN'
            else:
                raise ValueError("Language1 must be KAN or HIN (or Kannada/Hindi)")

            if User.objects.filter(username=uucms_id).exists():
                raise ValueError(f"UUCMS ID '{uucms_id}' already exists")
            if StudentProfile.objects.filter(roll_no=uucms_id).exists():
                raise ValueError(f"Roll number '{uucms_id}' already exists")

            # Find course – try code, then name, then code from first part of split
            course = Course.objects.filter(code=course_code).first()
            if not course:
                course = Course.objects.filter(name__iexact=course_code).first()
            if not course and '-' in course_code:
                parts = course_code.split('-')
                if parts[0].strip():
                    course = Course.objects.filter(code=parts[0].strip()).first()
            if not course:
                raise ValueError(f"Course '{course_code}' not found")

            section = Section.objects.filter(course=course, semester=semester, name=section_name).first()
            if not section:
                raise ValueError(f"Section '{section_name}' not found for course {course.code} semester {semester}")

            with transaction.atomic():
                user = User.objects.create_user(
                    username=uucms_id,
                    password='student123',
                    first_name=full_name,
                    last_name='',
                    email=email,
                    phone_number=phone,
                    user_type='student'
                )
                profile = StudentProfile.objects.create(
                    user=user,
                    roll_no=uucms_id,
                    department=course.department,
                    registration_complete=False
                )
                enrollment = StudentEnrollment.objects.create(
                    student=profile,
                    course=course,
                    semester=semester,
                    section=section
                )
                from attendance.utils import get_mandatory_subjects, get_language_subject, get_english_subject
                core = get_mandatory_subjects(course, semester, section)
                lang1 = get_language_subject(course, semester, language1)
                lang2 = get_english_subject(course, semester)
                all_subjects = list(core) + [lang1, lang2]
                enrollment.subjects.add(*all_subjects)
                enrollment.save()
            results['success'] += 1
            print(f"Student {uucms_id} created successfully")
        except Exception as e:
            print(f"Student row ERROR: {str(e)}")
            raise

    elif user_type == 'lecturer':
        try:
            if len(row) < 6:
                raise ValueError(f"Insufficient columns. Expected 6, got {len(row)}")
            full_name, username, email, phone, designation, course_code = row[:6]
            full_name = clean_text(full_name)
            username = clean_text(username)
            email = clean_text(email)
            phone = clean_text(phone)
            designation = clean_text(designation)
            course_code = clean_text(course_code)

            print(f"Lecturer parsed: full_name={full_name}, username={username}, email={email}, designation={designation}, course_code={course_code}")

            if not all([full_name, username, email, designation]):
                raise ValueError("Missing required fields")
            if User.objects.filter(username=username).exists():
                raise ValueError(f"Username '{username}' already exists")

            course = None
            if course_code:
                course = Course.objects.filter(code=course_code).first()
                if not course:
                    raise ValueError(f"Assigned course '{course_code}' not found")

            with transaction.atomic():
                user = User.objects.create_user(
                    username=username,
                    password='staff123',
                    first_name=full_name,
                    last_name='',
                    email=email,
                    phone_number=phone,
                    user_type='lecturer',
                    assigned_course=course
                )
                StaffProfile.objects.create(
                    user=user,
                    employee_id=username,
                    department=course.name if course else '',
                    designation=designation,
                    registration_complete=False
                )
            results['success'] += 1
            print(f"Lecturer {username} created successfully")
        except Exception as e:
            logger.error(f"Lecturer row error: {str(e)}")
            print(f"Lecturer row ERROR: {str(e)}")
            raise

    elif user_type == 'staff':
        try:
            if len(row) < 5:
                raise ValueError(f"Insufficient columns. Expected 5, got {len(row)}")
            full_name, username, email, phone, designation = row[:5]
            full_name = clean_text(full_name)
            username = clean_text(username)
            email = clean_text(email)
            phone = clean_text(phone)
            designation = clean_text(designation)

            print(f"Staff parsed: full_name={full_name}, username={username}, email={email}, designation={designation}")

            if not all([full_name, username, email, designation]):
                raise ValueError("Missing required fields")
            if User.objects.filter(username=username).exists():
                raise ValueError(f"Username '{username}' already exists")

            with transaction.atomic():
                user = User.objects.create_user(
                    username=username,
                    password='staff123',
                    first_name=full_name,
                    last_name='',
                    email=email,
                    phone_number=phone,
                    user_type='staff'
                )
                StaffProfile.objects.create(
                    user=user,
                    employee_id=username,
                    department='',
                    designation=designation,
                    registration_complete=False
                )
            results['success'] += 1
            print(f"Staff {username} created successfully")
        except Exception as e:
            logger.error(f"Staff row error: {str(e)}")
            print(f"Staff row ERROR: {str(e)}")
            raise

    else:
        raise ValueError(f"Unsupported user type: {user_type}")


def redirect_to_dashboard(request, user):
    """Redirect user to their respective dashboard based on user type."""
    dashboard_map = {
        'principal': 'principal_dashboard',
        'hod': 'hod_dashboard',
        'lecturer': 'lecturer_dashboard',
        'staff': 'staff_dashboard',
        'admin': 'admin_dashboard',
    }
    if user.user_type in dashboard_map:
        return redirect(dashboard_map[user.user_type])

    if user.user_type == 'student':
        try:
            profile = user.student_profile
            if not profile.registration_complete:
                messages.warning(request, 'Please complete your face registration.')
                return redirect('face_registration')
        except StudentProfile.DoesNotExist:
            messages.warning(request, 'Please complete your profile setup.')
            return redirect('face_registration')
        return redirect('student_dashboard')

    return redirect('home')


# =============================================================================
# Authentication views
# =============================================================================

def login_view(request):
    if request.user.is_authenticated:
        return redirect_to_dashboard(request, request.user)

    if request.method == 'POST':
        username = request.POST.get('username', '').strip()
        password = request.POST.get('password', '')
        try:
            user_obj = User.objects.get(username__iexact=username)
            actual_username = user_obj.username
        except User.DoesNotExist:
            actual_username = username

        user = authenticate(request, username=actual_username, password=password)
        if user is not None:
            login(request, user)
            messages.success(request, f'Welcome back, {user.get_full_name() or user.username}!')
            return redirect_to_dashboard(request, user)
        else:
            messages.error(request, 'Invalid username or password.')

    return render(request, 'accounts/login.html')


def logout_view(request):
    logout(request)
    messages.success(request, 'You have been logged out successfully.')
    return redirect('login')


def register_view(request):
    if request.user.is_authenticated:
        return redirect('home')

    if request.method == 'POST':
        user_type = request.POST.get('user_type')
        username = request.POST.get('username', '').strip()
        email = request.POST.get('email', '').strip()
        password = request.POST.get('password')
        confirm_password = request.POST.get('confirm_password')
        first_name = request.POST.get('first_name', '').strip()
        last_name = request.POST.get('last_name', '').strip()
        phone_number = request.POST.get('phone_number', '').strip()

        if password != confirm_password:
            messages.error(request, 'Passwords do not match.')
            return render(request, 'accounts/register.html')

        if User.objects.filter(username=username).exists():
            messages.error(request, 'Username already exists.')
            return render(request, 'accounts/register.html')

        if User.objects.filter(email=email).exists():
            messages.error(request, 'Email already registered.')
            return render(request, 'accounts/register.html')

        try:
            user = User.objects.create_user(
                username=username,
                email=email,
                password=password,
                first_name=first_name,
                last_name=last_name,
                user_type=user_type,
                phone_number=phone_number
            )

            if user_type == 'student':
                roll_no = request.POST.get('roll_no', '').strip()
                department = request.POST.get('department', '').strip()
                if StudentProfile.objects.filter(roll_no=roll_no).exists():
                    user.delete()
                    messages.error(request, 'Roll number already exists.')
                    return render(request, 'accounts/register.html')
                StudentProfile.objects.create(
                    user=user,
                    roll_no=roll_no,
                    department=department
                )
                messages.success(request, 'Registration successful! Please complete your face registration.')
                login(request, user)
                return redirect('face_registration')
            else:
                messages.success(request, 'Registration successful! Please login.')
                return redirect('login')

        except Exception as e:
            messages.error(request, f'Registration failed: {str(e)}')

    return render(request, 'accounts/register.html')


# =============================================================================
# Face registration and profile views
# =============================================================================

@login_required
def face_registration_view(request):
    if request.user.user_type != 'student':
        messages.error(request, 'Only students can register face.')
        return redirect('home')

    student_profile = get_object_or_404(StudentProfile, user=request.user)

    if request.method == 'POST':
        face_image = request.FILES.get('face_image')
        if face_image:
            success, msg = handle_face_image(student_profile, face_image)
            if success:
                messages.success(request, msg)
                return redirect('student_dashboard')
            else:
                messages.error(request, msg)
        else:
            face_data = request.POST.get('face_data')
            if face_data:
                try:
                    if ',' in face_data:
                        face_data = face_data.split(',')[1]
                    img_bytes = base64.b64decode(face_data)
                    np_arr = np.frombuffer(img_bytes, np.uint8)
                    img = cv2.imdecode(np_arr, cv2.IMREAD_COLOR)
                    rgb_img = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
                    pil_img = Image.fromarray(rgb_img)
                    img_io = io.BytesIO()
                    pil_img.save(img_io, format='JPEG')
                    img_io.seek(0)

                    success, msg = handle_face_image(student_profile, img_io)
                    if success:
                        return JsonResponse({'success': True, 'message': msg})
                    else:
                        return JsonResponse({'success': False, 'message': msg})
                except Exception as e:
                    return JsonResponse({'success': False, 'message': f'Error: {str(e)}'})

    return render(request, 'accounts/face_registration.html', {'student_profile': student_profile})


@login_required
@csrf_exempt
def update_face_encoding(request):
    if request.user.user_type != 'student':
        return JsonResponse({'error': 'Unauthorized'}, status=403)

    if request.method != 'POST':
        return JsonResponse({'error': 'Invalid method'}, status=405)

    try:
        data = json.loads(request.body)
        face_data = data.get('face_data')
        if not face_data:
            return JsonResponse({'success': False, 'message': 'No face data provided.'})

        if ',' in face_data:
            face_data = face_data.split(',')[1]

        img_bytes = base64.b64decode(face_data)
        np_arr = np.frombuffer(img_bytes, np.uint8)
        img = cv2.imdecode(np_arr, cv2.IMREAD_COLOR)
        rgb_img = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
        face_encodings = face_recognition.face_encodings(rgb_img)

        if not face_encodings:
            return JsonResponse({'success': False, 'message': 'No face detected.'})

        student_profile = get_object_or_404(StudentProfile, user=request.user)
        student_profile.set_face_encoding(face_encodings[0])
        student_profile.save()
        return JsonResponse({'success': True, 'message': 'Face encoding updated.'})

    except Exception as e:
        return JsonResponse({'success': False, 'message': f'Error: {str(e)}'})


@login_required
def profile_view(request):
    context = {}
    if request.user.user_type == 'student':
        try:
            context['student_profile'] = StudentProfile.objects.get(user=request.user)
        except StudentProfile.DoesNotExist:
            pass
    return render(request, 'accounts/profile.html', context)


@login_required
def edit_profile(request):
    if request.method == 'POST':
        user = request.user
        user.first_name = request.POST.get('first_name', user.first_name).strip()
        user.last_name = request.POST.get('last_name', user.last_name).strip()
        user.email = request.POST.get('email', user.email).strip()
        user.phone_number = request.POST.get('phone_number', user.phone_number).strip()
        user.save()

        if user.user_type == 'student':
            try:
                student_profile = StudentProfile.objects.get(user=user)
                student_profile.department = request.POST.get('department', student_profile.department).strip()
                student_profile.save()
            except StudentProfile.DoesNotExist:
                pass

        messages.success(request, 'Profile updated successfully!')
        return redirect('profile')

    context = {}
    if request.user.user_type == 'student':
        try:
            context['student_profile'] = StudentProfile.objects.get(user=request.user)
        except StudentProfile.DoesNotExist:
            pass
    return render(request, 'accounts/edit_profile.html', context)


@login_required
def change_password(request):
    if request.method == 'POST':
        user = request.user
        old_password = request.POST.get('old_password')
        new_password = request.POST.get('new_password')
        confirm_password = request.POST.get('confirm_password')

        if not user.check_password(old_password):
            messages.error(request, 'Current password is incorrect.')
            return redirect('change_password')

        if new_password != confirm_password:
            messages.error(request, 'New passwords do not match.')
            return redirect('change_password')

        if len(new_password) < 8:
            messages.error(request, 'Password must be at least 8 characters long.')
            return redirect('change_password')

        user.set_password(new_password)
        user.save()
        login(request, user)
        messages.success(request, 'Password changed successfully!')
        return redirect('profile')

    return render(request, 'accounts/change_password.html')


# =============================================================================
# Lecturer views
# =============================================================================

@login_required
def student_list(request):
    if request.user.user_type != 'lecturer':
        messages.error(request, 'Access denied.')
        return redirect('home')

    students = StudentProfile.objects.select_related('user').all()
    return render(request, 'accounts/student_list.html', {'students': students})


@login_required
def student_detail(request, student_id):
    if request.user.user_type != 'lecturer':
        messages.error(request, 'Access denied.')
        return redirect('home')

    student = get_object_or_404(StudentProfile.objects.select_related('user'), id=student_id)

    total_classes = AttendanceRecord.objects.filter(student=student).count()
    present_count = AttendanceRecord.objects.filter(student=student, status='Present').count()
    attendance_percentage = (present_count / total_classes * 100) if total_classes > 0 else 0

    context = {
        'student': student,
        'total_classes': total_classes,
        'present_count': present_count,
        'attendance_percentage': round(attendance_percentage, 2),
    }
    return render(request, 'accounts/student_detail.html', context)


# =============================================================================
# Admin views
# =============================================================================

@login_required
def admin_dashboard(request):
    if request.user.user_type != 'admin':
        messages.error(request, "Access denied.")
        return redirect('dashboard')

    context = {
        'total_students': User.objects.filter(user_type='student').count(),
        'total_teaching': User.objects.filter(user_type='lecturer').count(),
        'total_nonteaching': User.objects.filter(user_type='staff').count(),
    }
    return render(request, 'admin/admin_dashboard.html', context)


@login_required
def admin_register_choice(request):
    if request.user.user_type != 'admin':
        messages.error(request, "Access denied.")
        return redirect('admin_dashboard')
    return render(request, 'admin/admin_register_choice.html')


@login_required
def admin_register_student(request):
    if request.user.user_type not in ['admin', 'principal']:
        messages.error(request, 'Access denied.')
        return redirect('home')

    # Confirmation POST
    if request.method == 'POST' and request.POST.get('confirm') == '1':
        temp_data = request.session.get('temp_student_data', {})
        if not temp_data:
            messages.error(request, 'Session expired. Please start over.')
            return redirect('admin_register_student')

        form = StudentRegistrationForm(temp_data)
        if form.is_valid():
            data = form.cleaned_data
            with transaction.atomic():
                user = User.objects.create_user(
                    username=data['uucms_id'],
                    password=data['password'],
                    first_name=data['student_name'],
                    last_name='',
                    email=data['email'],
                    phone_number=data.get('phone_number', ''),
                    user_type='student'
                )
                student_profile = StudentProfile.objects.create(
                    user=user,
                    roll_no=data['uucms_id'],
                    department=data['course'].department,
                    registration_complete=False
                )

                # Handle face image
                face_data = request.session.get('temp_face_image')
                if face_data:
                    try:
                        from django.core.files.base import ContentFile
                        import base64
                        image_data = base64.b64decode(face_data)
                        file_name = request.session.get('temp_face_image_name', 'face.jpg')
                        student_profile.face_image.save(file_name, ContentFile(image_data), save=True)
                        from attendance.utils import encode_face_from_image
                        encoding = encode_face_from_image(student_profile.face_image.path)
                        if encoding is not None:
                            student_profile.set_face_encoding(encoding)
                            student_profile.registration_complete = True
                            student_profile.save()
                        else:
                            messages.warning(request, 'Face image saved but no face detected. Please re-upload.')
                    except Exception as e:
                        messages.error(request, f'Error processing face image: {str(e)}')

                # ✅ Create enrollment – no 'language1' here
                enrollment = StudentEnrollment.objects.create(
                    student=student_profile,
                    course=data['course'],
                    semester=data['semester'],
                    section=data['section']
                )

                # Assign subjects using the language choice
                core_subjects = get_mandatory_subjects(data['course'], data['semester'], data['section'])
                lang_subject = get_language_subject(data['course'], data['semester'], data['language1'])
                english_subject = get_english_subject(data['course'], data['semester'])
                all_subjects = list(core_subjects) + [lang_subject, english_subject]
                enrollment.subjects.add(*all_subjects)
                enrollment.save()

            request.session.pop('temp_student_data', None)
            request.session.pop('temp_face_image', None)
            request.session.pop('temp_face_image_name', None)

            messages.success(request, f'Student {user.get_full_name()} registered successfully!')
            return redirect('admin_dashboard')
        else:
            messages.error(request, 'Invalid data. Please try again.')
            return redirect('admin_register_student')

    # First POST (show confirmation)
    if request.method == 'POST':
        form = StudentRegistrationForm(request.POST, request.FILES)
        if form.is_valid():
            data = form.cleaned_data
            request.session['temp_student_data'] = {
                'uucms_id': data['uucms_id'],
                'student_name': data['student_name'],
                'email': data['email'],
                'phone_number': data.get('phone_number', ''),
                'password': data['password'],
                'course': data['course'].id,
                'semester': data['semester'],
                'section': data['section'].id,
                'language1': data['language1'],
            }

            if request.FILES.get('face_image'):
                img = request.FILES['face_image']
                import base64
                img.seek(0)  # reset pointer
                encoded = base64.b64encode(img.read()).decode('utf-8')
                request.session['temp_face_image'] = encoded
                request.session['temp_face_image_name'] = img.name

            core_subjects = get_mandatory_subjects(data['course'], data['semester'], data['section'])
            lang_subject = get_language_subject(data['course'], data['semester'], data['language1'])
            english_subject = get_english_subject(data['course'], data['semester'])
            all_subjects = list(core_subjects) + [lang_subject, english_subject]
            subject_names = [f"{s.code} - {s.name}" for s in all_subjects]

            confirm_data = {
                'full_name': data['student_name'],
                'uucms_id': data['uucms_id'],
                'email': data['email'],
                'phone': data.get('phone_number', ''),
                'course': data['course'],
                'semester': data['semester'],
                'section': data['section'],
                'language1': 'Kannada' if data['language1'] == 'KAN' else 'Hindi',
                'language2': 'Basic English',
                'subjects': subject_names,
            }
            context = {
                'form': form,
                'confirm_data': confirm_data,
                'show_confirm': True,
            }
            return render(request, 'admin/register_student.html', context)
        else:
            context = {'form': form, 'show_confirm': False}
            return render(request, 'admin/register_student.html', context)

    # GET
    form = StudentRegistrationForm()
    context = {'form': form, 'show_confirm': False}
    return render(request, 'admin/register_student.html', context)

@login_required
def admin_register_teaching(request):
    if request.user.user_type != 'admin':
        messages.error(request, "Access denied.")
        return redirect('admin_dashboard')

    if request.method == 'POST':
        form = TeachingStaffRegistrationForm(request.POST, request.FILES)
        if form.is_valid():
            data = form.cleaned_data
            with transaction.atomic():
                user = User.objects.create_user(
                    username=data['username'],
                    password=data['password'],
                    first_name=data['full_name'],
                    last_name='',
                    email=data['email'],
                    phone_number=data.get('phone_number', ''),
                    user_type='lecturer',
                    assigned_course=None
                )
                staff_profile = StaffProfile.objects.create(
                    user=user,
                    employee_id=data['username'],
                    department=data['course'].name,
                    designation=data['designation'],
                    registration_complete=False
                )

                if data.get('face_image'):
                    success, msg = handle_face_image(staff_profile, data['face_image'])
                    if not success:
                        messages.warning(request, f"Face registration issue: {msg}")

                messages.success(request, f"Teaching staff {user.get_full_name()} registered.")
                return redirect('admin_teaching_list')
        else:
            # re-render with errors
            return render(request, 'admin/admin_register_teaching.html', {'form': form})
    else:
        form = TeachingStaffRegistrationForm()
        return render(request, 'admin/admin_register_teaching.html', {'form': form})


@login_required
def admin_register_nonteaching(request):
    if request.user.user_type != 'admin':
        messages.error(request, "Access denied.")
        return redirect('admin_dashboard')

    if request.method == 'POST':
        form = NonTeachingStaffRegistrationForm(request.POST, request.FILES)
        if form.is_valid():
            data = form.cleaned_data
            with transaction.atomic():
                user = User.objects.create_user(
                    username=data['username'],
                    password=data['password'],
                    first_name=data['full_name'],
                    last_name='',
                    email=data['email'],
                    phone_number=data.get('phone_number', ''),
                    user_type='staff',
                    assigned_course=None
                )
                staff_profile = StaffProfile.objects.create(
                    user=user,
                    employee_id=data['username'],
                    department='',
                    designation=data['designation'],
                    registration_complete=False
                )

                if data.get('face_image'):
                    success, msg = handle_face_image(staff_profile, data['face_image'])
                    if not success:
                        messages.warning(request, f"Face registration issue: {msg}")

                messages.success(request, f"Non-teaching staff {user.get_full_name()} registered.")
                return redirect('admin_nonteaching_list')
        else:
            return render(request, 'admin/admin_register_nonteaching.html', {'form': form})
    else:
        form = NonTeachingStaffRegistrationForm()
        return render(request, 'admin/admin_register_nonteaching.html', {'form': form})


@login_required
def admin_student_list(request):
    if request.user.user_type != 'admin':
        messages.error(request, "Access denied.")
        return redirect('admin_dashboard')

    enrollments = StudentEnrollment.objects.select_related(
        'student__user', 'course', 'section'
    ).all()

    search = request.GET.get('search', '')
    course_id = request.GET.get('course')
    year = request.GET.get('year')
    section_name = request.GET.get('section')

    if search:
        enrollments = enrollments.filter(
            Q(student__user__first_name__icontains=search) |
            Q(student__user__last_name__icontains=search) |
            Q(student__user__username__icontains=search) |
            Q(student__roll_no__icontains=search)
        )
    if course_id:
        enrollments = enrollments.filter(course_id=course_id)
    if year:
        enrollments = enrollments.filter(semester=year)
    if section_name:
        enrollments = enrollments.filter(section__name=section_name)

    courses = Course.objects.all()
    sections = Section.objects.values_list('name', flat=True).distinct().order_by('name')

    context = {
        'enrollments': enrollments,
        'courses': courses,
        'sections': sections,
        'search': search,
        'selected_course': course_id,
        'selected_year': year,
        'selected_section': section_name,
    }
    return render(request, 'admin/admin_student_list.html', context)


@login_required
def admin_user_detail(request, user_id):
    if request.user.user_type != 'admin':
        messages.error(request, "Access denied.")
        return redirect('admin_dashboard')

    user_obj = get_object_or_404(User, id=user_id)
    profile = None
    if user_obj.user_type == 'student':
        try:
            profile = user_obj.student_profile
        except StudentProfile.DoesNotExist:
            pass

    if request.method == 'POST':
        if 'delete' in request.POST:
            user_obj.delete()
            messages.success(request, "User deleted.")
            return redirect('admin_dashboard')

        form = AdminUserUpdateForm(request.POST, instance=user_obj)
        if form.is_valid():
            form.save()
            messages.success(request, "User updated.")

            # Handle face image
            face_image = request.FILES.get('face_image')
            if face_image:
                if user_obj.user_type == 'student' and profile:
                    success, msg = handle_face_image(profile, face_image)
                    if success:
                        messages.success(request, "Face image updated.")
                    else:
                        messages.warning(request, f"Face update issue: {msg}")
                elif user_obj.user_type in ['lecturer', 'hod', 'staff']:
                    staff_profile = getattr(user_obj, 'staff_profile', None)
                    if staff_profile:
                        success, msg = handle_face_image(staff_profile, face_image)
                        if success:
                            messages.success(request, "Face image updated.")
                        else:
                            messages.warning(request, f"Face update issue: {msg}")
                    else:
                        messages.warning(request, "Staff profile not found.")
                else:
                    messages.warning(request, "Face image not applicable for this user type.")

            return redirect('admin_user_detail', user_id=user_obj.id)
        else:
            messages.error(request, "Please correct the errors below.")
    else:
        form = AdminUserUpdateForm(instance=user_obj)

    context = {
        'user_obj': user_obj,
        'profile': profile,
        'form': form,
    }
    return render(request, 'admin/admin_user_detail.html', context)


@login_required
def admin_teaching_list(request):
    if request.user.user_type != 'admin':
        messages.error(request, "Access denied.")
        return redirect('admin_dashboard')

    staff = User.objects.filter(user_type='lecturer')
    search = request.GET.get('search', '')
    course_id = request.GET.get('course')
    department = request.GET.get('department')
    designation = request.GET.get('designation')

    if search:
        staff = staff.filter(
            Q(first_name__icontains=search) |
            Q(last_name__icontains=search) |
            Q(username__icontains=search) |
            Q(email__icontains=search)
        )
    if course_id:
        staff = staff.filter(assigned_course_id=course_id)
    if department:
        staff = staff.filter(staff_profile__department__icontains=department)
    if designation:
        staff = staff.filter(staff_profile__designation=designation)

    courses = Course.objects.all()
    departments = StaffProfile.objects.filter(
        user__user_type='lecturer'
    ).values_list('department', flat=True).distinct().exclude(department='')

    TEACHING_DESIGNATIONS = ['Professor', 'Associate Professor', 'Assistant Professor', 'Lecturer/Instructor']
    existing = StaffProfile.objects.filter(
        user__user_type='lecturer'
    ).values_list('designation', flat=True).distinct().exclude(designation='')
    designations = TEACHING_DESIGNATIONS.copy()
    for d in existing:
        if d not in designations:
            designations.append(d)

    context = {
        'staff': staff,
        'search': search,
        'courses': courses,
        'selected_course': course_id,
        'departments': departments,
        'selected_department': department,
        'designations': designations,
        'selected_designation': designation,
    }
    return render(request, 'admin/admin_teaching_staff_list.html', context)


@login_required
def admin_nonteaching_list(request):
    if request.user.user_type != 'admin':
        messages.error(request, "Access denied.")
        return redirect('admin_dashboard')

    staff = User.objects.filter(user_type='staff')
    search = request.GET.get('search', '')
    designation = request.GET.get('designation')

    if search:
        staff = staff.filter(
            Q(first_name__icontains=search) |
            Q(last_name__icontains=search) |
            Q(username__icontains=search) |
            Q(email__icontains=search)
        )
    if designation:
        staff = staff.filter(staff_profile__designation=designation)

    NON_TEACHING_DESIGNATIONS = ['FDA', 'SDA', 'Peon', 'Lab Attender', 'Sweeper']

    context = {
        'staff': staff,
        'search': search,
        'type': 'Non-Teaching',
        'designations': NON_TEACHING_DESIGNATIONS,
        'selected_designation': designation,
    }
    return render(request, 'admin/admin_nonteaching_staff_list.html', context)


import logging
logger = logging.getLogger(__name__)

@login_required
def admin_bulk_register(request):
    if request.user.user_type != 'admin':
        messages.error(request, "Access denied.")
        return redirect('admin_dashboard')

    if request.method == 'POST':
        form = BulkUploadForm(request.POST, request.FILES)
        if form.is_valid():
            user_type = form.cleaned_data['user_type']
            excel_file = request.FILES['excel_file']
            logger.info(f"=== BULK REGISTRATION START: user_type={user_type}, file={excel_file.name} ===")
            print(f"=== BULK REGISTRATION START: user_type={user_type}, file={excel_file.name} ===")
            try:
                # Validate file extension
                if not excel_file.name.endswith(('.xlsx', '.xls')):
                    error_msg = "File must be an Excel file (.xlsx or .xls)."
                    logger.error(error_msg)
                    messages.error(request, error_msg)
                    return render(request, 'admin/admin_bulk_register.html', {'form': form})

                wb = load_workbook(excel_file)
                ws = wb.active
                rows = list(ws.iter_rows(values_only=True))
                logger.info(f"Total rows (including header): {len(rows)}")
                print(f"Total rows (including header): {len(rows)}")
                if len(rows) < 2:
                    error_msg = "File is empty or missing data rows."
                    logger.error(error_msg)
                    messages.error(request, error_msg)
                    return render(request, 'admin/admin_bulk_register.html', {'form': form})

                headers = rows[0]
                data = rows[1:]
                logger.info(f"Headers: {headers}")
                print(f"Headers: {headers}")
                logger.info(f"Data rows count: {len(data)}")
                print(f"Data rows count: {len(data)}")

                expected = get_expected_headers(user_type)
                headers_clean = [str(h).strip() for h in headers]
                expected_clean = [e.strip() for e in expected]
                logger.info(f"Cleaned headers: {headers_clean}")
                logger.info(f"Expected headers: {expected_clean}")
                print(f"Cleaned headers: {headers_clean}")
                print(f"Expected headers: {expected_clean}")

                if headers_clean != expected_clean:
                    error_msg = f"Invalid headers. Expected: {', '.join(expected)}. Got: {', '.join(headers_clean)}"
                    logger.error(error_msg)
                    messages.error(request, error_msg)
                    return render(request, 'admin/admin_bulk_register.html', {'form': form})

                results = {'success': 0, 'failed': 0, 'errors': []}
                for row_idx, row in enumerate(data, start=2):
                    # Skip completely empty rows
                    if not any(cell is not None and str(cell).strip() for cell in row):
                        logger.info(f"Skipping empty row {row_idx}")
                        continue
                    logger.info(f"Processing row {row_idx}: {row}")
                    print(f"Processing row {row_idx}: {row}")
                    try:
                        process_bulk_row(user_type, row, results)
                    except Exception as e:
                        results['failed'] += 1
                        error_detail = f"Row {row_idx}: {str(e)}"
                        results['errors'].append(error_detail)
                        logger.error(error_detail)
                        print(f"ERROR: {error_detail}")

                logger.info(f"Bulk registration results: Success={results['success']}, Failed={results['failed']}")
                print(f"Bulk registration results: Success={results['success']}, Failed={results['failed']}")

                messages.success(
                    request,
                    f"Bulk registration completed. Success: {results['success']}, Failed: {results['failed']}."
                )
                if results['errors']:
                    errors_str = '; '.join(results['errors'][:10])
                    if len(results['errors']) > 10:
                        errors_str += f" and {len(results['errors'])-10} more."
                    messages.warning(request, f"Errors: {errors_str}")
                return redirect('admin_dashboard')
            except Exception as e:
                error_msg = f"Error processing file: {str(e)}"
                logger.exception(error_msg)
                print(f"EXCEPTION: {error_msg}")
                import traceback
                traceback.print_exc()
                messages.error(request, error_msg)
        else:
            logger.warning("Bulk upload form invalid")
            messages.error(request, "Invalid form data.")
    else:
        form = BulkUploadForm()

    return render(request, 'admin/admin_bulk_register.html', {'form': form})


@login_required
def admin_download_sample(request, user_type):
    if request.user.user_type != 'admin':
        messages.error(request, "Access denied.")
        return redirect('admin_dashboard')

    headers = get_expected_headers(user_type)
    if not headers:
        return HttpResponse("Invalid user type", status=400)

    wb = Workbook()
    ws = wb.active
    ws.title = "Sample Data"

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = min(len(header) + 4, 30)

    # No sample data row – only headers
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=sample_{user_type}.xlsx'
    wb.save(response)
    return response
