import io
import json
import calendar
from datetime import datetime, timedelta

import openpyxl
from django.contrib import messages
from django.contrib.auth import get_user_model
from django.contrib.auth.decorators import login_required
from django.db.models import Count, Q
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.template.loader import get_template
from django.utils import timezone
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from xhtml2pdf import pisa

from accounts.decorators import admin_required
from accounts.models import Course, Section, StaffProfile, StudentEnrollment, StudentProfile, Subject, User

from .models import AttendanceRecord, AttendanceSession, HolidaySetting, StaffAttendance, SystemSetting
from attendance.utils import (
    get_mandatory_subjects,
    get_language_subject,
    get_english_subject,
    # ... other functions
)
from accounts.views import get_mandatory_subjects
User = get_user_model()


# =============================================================================
# HOME & DASHBOARDS
# =============================================================================

@login_required
def home(request):
    """Redirect based on user type after login."""
    if request.user.user_type == 'principal':
        return redirect('principal_dashboard')
    elif request.user.user_type == 'hod':
        return redirect('hod_dashboard')
    elif request.user.user_type == 'lecturer':
        return redirect('lecturer_dashboard')
    elif request.user.user_type == 'student':
        return redirect('student_dashboard')
    else:
        return redirect('profile')


@login_required
def lecturer_dashboard(request):
    """Lecturer dashboard showing active and past sessions."""
    if request.user.user_type != 'lecturer':
        return redirect('student_dashboard')

    active_sessions = AttendanceSession.objects.filter(
        lecturer=request.user,
        is_active=True
    ).select_related('course', 'section', 'subject')

    past_sessions = AttendanceSession.objects.filter(
        lecturer=request.user,
        is_active=False
    ).select_related('course', 'section', 'subject')[:20]

    context = {
        'active_sessions': active_sessions,
        'past_sessions': past_sessions,
    }
    return render(request, 'attendance/lecturer_dashboard.html', context)


@login_required
def student_dashboard(request):
    if request.user.user_type != 'student':
        return redirect('lecturer_dashboard')

    student_profile = get_object_or_404(StudentProfile, user=request.user)

    enrollments = StudentEnrollment.objects.filter(
        student=student_profile
    ).select_related('course', 'section').prefetch_related('subjects')

    q_filter = Q()
    for enrollment in enrollments:
        q_filter |= Q(
            course=enrollment.course,
            semester=enrollment.semester,
            subject__in=enrollment.subjects.all(),
        ) & (Q(section=enrollment.section) | Q(section__isnull=True))

    active_sessions = AttendanceSession.objects.filter(
        q_filter,
        is_active=True
    ).select_related('course', 'section', 'subject', 'lecturer').distinct()

    attendance_history = AttendanceRecord.objects.filter(
        student=student_profile
    ).select_related('session__subject', 'session__course').order_by('-timestamp')[:20]

    context = {
        'student_profile': student_profile,
        'active_sessions': active_sessions,
        'enrollments': enrollments,
        'attendance_history': attendance_history,
    }
    return render(request, 'attendance/student_dashboard.html', context)


@login_required
def hod_dashboard(request):
    if request.user.user_type != 'hod':
        return redirect('home')

    assigned_course = request.user.assigned_course
    if not assigned_course:
        messages.error(request, 'No course assigned to you. Contact administrator.')
        return redirect('home')

    course = assigned_course

    students_count = StudentEnrollment.objects.filter(course=course).values('student').distinct().count()
    lecturers_count = User.objects.filter(user_type='lecturer', assigned_course=course).count()
    active_sessions = AttendanceSession.objects.filter(course=course, is_active=True)
    recent_sessions = AttendanceSession.objects.filter(
        course=course
    ).select_related('lecturer', 'subject', 'section').order_by('-created_at')[:10]

    year_data = []
    for year in range(1, course.duration_years + 1):
        sem_start = (year - 1) * 2 + 1
        sem_end = year * 2

        sections = Section.objects.filter(
            course=course,
            semester__gte=sem_start,
            semester__lte=sem_end
        ).distinct()

        section_data = []
        total_students = 0

        for section in sections:
            count = StudentEnrollment.objects.filter(
                course=course,
                semester__gte=sem_start,
                semester__lte=sem_end,
                section=section
            ).values('student').distinct().count()

            section_data.append({
                'section': section,
                'count': count,
            })
            total_students += count

        year_data.append({
            'year': year,
            'sections': section_data,
            'total': total_students,
        })

    context = {
        'course': course,
        'students_count': students_count,
        'lecturers_count': lecturers_count,
        'active_sessions': active_sessions,
        'recent_sessions': recent_sessions,
        'year_data': year_data,
    }
    return render(request, 'attendance/hod_dashboard.html', context)


# =============================================================================
# PRINCIPAL VIEWS
# =============================================================================

@login_required
def principal_dashboard(request):
    if request.user.user_type != 'principal':
        return redirect('home')

    teaching_staff_count = User.objects.filter(user_type__in=['lecturer', 'hod']).count()
    non_teaching_staff_count = User.objects.filter(user_type='staff').count()
    students_count = StudentProfile.objects.count()

    context = {
        'teaching_staff_count': teaching_staff_count,
        'non_teaching_staff_count': non_teaching_staff_count,
        'students_count': students_count,
    }
    return render(request, 'attendance/principal_dashboard.html', context)


@login_required
def principal_course_types(request, type):
    if request.user.user_type != 'principal':
        return redirect('home')

    if type == 'students':
        ug_courses = Course.objects.filter(course_type='UG')
        pg_courses = Course.objects.filter(course_type='PG')
        context = {'type': type, 'ug_courses': ug_courses, 'pg_courses': pg_courses}
        return render(request, 'attendance/principal_students.html', context)

    return redirect('principal_dashboard')


@login_required
def principal_course_years(request, course_id):
    if request.user.user_type not in ['principal', 'hod']:
        messages.error(request, 'Access denied.')
        return redirect('home')

    course = get_object_or_404(Course, id=course_id)

    if request.user.user_type == 'hod':
        if not request.user.assigned_course or request.user.assigned_course.id != course.id:
            messages.error(request, 'Access denied. You can only view your assigned course.')
            return redirect('hod_dashboard')

    semesters = [
        {'number': 1, 'year': 1, 'label': 'First Year'},
        {'number': 3, 'year': 2, 'label': 'Second Year'},
        {'number': 5, 'year': 3, 'label': 'Third Year'},
    ]
    context = {'course': course, 'semesters': semesters}
    return render(request, 'attendance/course_years.html', context)


@login_required
def course_year_sections(request, course_id, year):
    if request.user.user_type not in ['principal', 'hod']:
        messages.error(request, 'Access denied.')
        return redirect('home')

    course = get_object_or_404(Course, id=course_id)

    if request.user.user_type == 'hod':
        if not request.user.assigned_course or request.user.assigned_course.id != course.id:
            messages.error(request, 'Access denied. You can only view your assigned course.')
            return redirect('hod_dashboard')

    sections = Section.objects.filter(course=course, semester=year)
    sections_data = []
    for section in sections:
        student_count = StudentEnrollment.objects.filter(
            course=course, semester=year, section=section
        ).count()
        sections_data.append({'section': section, 'student_count': student_count})

    context = {
        'course': course,
        'year': year,
        'sections_data': sections_data,
    }
    return render(request, 'attendance/section_list.html', context)


@login_required
def section_students(request, course_id, year, section_id):
    if request.user.user_type not in ['principal', 'hod']:
        messages.error(request, 'Access denied.')
        return redirect('home')

    course = get_object_or_404(Course, id=course_id)
    section = get_object_or_404(Section, id=section_id)

    if request.user.user_type == 'hod':
        if not request.user.assigned_course or request.user.assigned_course.id != course.id:
            messages.error(request, 'Access denied. You can only view your assigned course.')
            return redirect('hod_dashboard')

    enrollments = StudentEnrollment.objects.filter(
        course=course, semester=year, section=section
    ).select_related('student__user').prefetch_related('subjects')

    subject_ids = enrollments.values_list('subjects', flat=True).distinct()
    subjects = Subject.objects.filter(id__in=subject_ids).order_by('code')

    students_data = []
    for i, enrollment in enumerate(enrollments, 1):
        student = enrollment.student
        subject_attendance = []
        total_present = 0
        total_classes = 0

        enrolled_subjects = enrollment.subjects.all()

        for subject in subjects:
            if subject not in enrolled_subjects:
                subject_attendance.append({
                    'subject': subject,
                    'enrolled': False,
                    'present': None,
                    'total': None,
                    'percentage': None,
                })
                continue

            sessions = AttendanceSession.objects.filter(
                course=course, semester=year, section=section, subject=subject
            )
            subject_total_classes = sessions.count()
            subject_present = AttendanceRecord.objects.filter(
                session__in=sessions, student=student, status='Present'
            ).count()

            if subject_total_classes > 0:
                percentage = round((subject_present / subject_total_classes) * 100, 1)
                subject_attendance.append({
                    'subject': subject,
                    'enrolled': True,
                    'present': subject_present,
                    'total': subject_total_classes,
                    'percentage': percentage,
                })
                total_present += subject_present
                total_classes += subject_total_classes
            else:
                subject_attendance.append({
                    'subject': subject,
                    'enrolled': True,
                    'present': 0,
                    'total': 0,
                    'percentage': 0,
                })

        overall_percentage = round((total_present / total_classes) * 100, 1) if total_classes > 0 else 0

        students_data.append({
            'sl_no': i,
            'student': student,
            'enrollment': enrollment,
            'subject_attendance': subject_attendance,
            'total_present': total_present,
            'total_classes': total_classes,
            'overall_percentage': overall_percentage,
        })

    export = request.GET.get('export')
    if export == 'excel':
        return export_students_excel(students_data, subjects, course, year, section)
    elif export == 'pdf':
        return export_students_pdf(students_data, subjects, course, year, section)

    context = {
        'course': course,
        'year': year,
        'section': section,
        'subjects': subjects,
        'students_data': students_data,
    }
    return render(request, 'attendance/student_list_attendance.html', context)


def export_students_excel(students_data, subjects, course, year, section):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Student Attendance"

    headers = ['Sl No', 'UUCMS ID', 'Student Name']
    for subject in subjects:
        headers.append(f"{subject.code} (P/T)")
        headers.append(f"{subject.code} %")
    headers.append('Overall %')
    headers.append('Total (P/T)')

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal='center')
        cell.fill = PatternFill(start_color="1a237e", end_color="1a237e", fill_type="solid")

    for row, data in enumerate(students_data, start=2):
        col = 1
        ws.cell(row=row, column=col, value=data['sl_no'])
        col += 1
        ws.cell(row=row, column=col, value=data['student'].roll_no)
        col += 1
        ws.cell(row=row, column=col, value=data['student'].user.get_full_name())
        col += 1

        for sub_att in data['subject_attendance']:
            ws.cell(row=row, column=col, value=f"{sub_att['present']}/{sub_att['total']}")
            col += 1
            ws.cell(row=row, column=col, value=sub_att['percentage'])
            col += 1

        ws.cell(row=row, column=col, value=data['overall_percentage'])
        col += 1
        ws.cell(row=row, column=col, value=f"{data['total_present']}/{data['total_classes']}")

    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except Exception:
                pass
        adjusted_width = min(max_length + 2, 30)
        ws.column_dimensions[col_letter].width = adjusted_width

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=students_{course.code}_Y{year}_Sec{section.name}.xlsx'
    wb.save(response)
    return response


def export_students_pdf(students_data, subjects, course, year, section):
    html = f"""
    <html>
    <head>
        <style>
            body {{ font-family: 'Helvetica', sans-serif; }}
            h2 {{ color: #1a237e; }}
            .header {{ margin-bottom: 20px; }}
            table {{ width: 100%; border-collapse: collapse; }}
            th {{ background-color: #1a237e; color: white; padding: 6px; text-align: left; font-size: 10px; }}
            td {{ padding: 4px 6px; border: 1px solid #ddd; font-size: 9px; }}
            tr:nth-child(even) {{ background-color: #f9f9f9; }}
            .text-center {{ text-align: center; }}
        </style>
    </head>
    <body>
        <div class="header">
            <h2>Student Attendance - {course.code} ({course.name})</h2>
            <p><strong>Year:</strong> {year} | <strong>Section:</strong> {section.name} | <strong>Total Students:</strong> {len(students_data)}</p>
        </div>
        <table>
            <thead>
                <tr>
                    <th>Sl No</th>
                    <th>UUCMS ID</th>
                    <th>Student Name</th>
    """

    for subject in subjects:
        html += f"<th>{subject.code}<br><small>P/T</small></th>"
        html += f"<th>{subject.code}<br><small>%</small></th>"

    html += """
                    <th>Overall %</th>
                    <th>Total<br><small>P/T</small></th>
                </tr>
            </thead>
            <tbody>
    """

    for data in students_data:
        html += f"""
                <tr>
                    <td class="text-center">{data['sl_no']}</td>
                    <td>{data['student'].roll_no}</td>
                    <td>{data['student'].user.get_full_name()}</td>
        """
        for sub_att in data['subject_attendance']:
            html += f"<td class='text-center'>{sub_att['present']}/{sub_att['total']}</td>"
            html += f"<td class='text-center'>{sub_att['percentage']}%</td>"

        html += f"""
                    <td class="text-center"><strong>{data['overall_percentage']}%</strong></td>
                    <td class="text-center">{data['total_present']}/{data['total_classes']}</td>
                </tr>
        """

    html += """
            </tbody>
        </table>
    </body>
    </html>
    """

    result = io.BytesIO()
    pdf = pisa.pisaDocument(io.BytesIO(html.encode("UTF-8")), result)
    if pdf.err:
        return HttpResponse('Error generating PDF', status=500)

    response = HttpResponse(result.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename=students_{course.code}_Y{year}_Sec{section.name}.pdf'
    return response


@login_required
def student_attendance_detail(request, student_id, course_id):
    if request.user.user_type not in ['principal', 'hod', 'lecturer']:
        messages.error(request, 'Access denied.')
        return redirect('home')

    student_profile = get_object_or_404(StudentProfile, id=student_id)
    course = get_object_or_404(Course, id=course_id)

    if request.user.user_type == 'hod':
        if not request.user.assigned_course or request.user.assigned_course.id != course.id:
            messages.error(request, 'Access denied. You can only view your assigned course students.')
            return redirect('hod_dashboard')

    if request.user.user_type == 'lecturer':
        if not request.user.assigned_course or request.user.assigned_course.id != course.id:
            messages.error(request, 'Access denied.')
            return redirect('lecturer_dashboard')

    enrollments = StudentEnrollment.objects.filter(student=student_profile, course=course).order_by('semester')

    attendance_details = []
    for enrollment in enrollments:
        subjects = enrollment.subjects.all()
        semester_data = {
            'semester': enrollment.semester,
            'section': enrollment.section,
            'subjects': []
        }

        for subject in subjects:
            sessions = AttendanceSession.objects.filter(
                course=course,
                semester=enrollment.semester,
                section=enrollment.section,
                subject=subject
            )
            total_classes = sessions.count()
            present_count = AttendanceRecord.objects.filter(
                session__in=sessions, student=student_profile, status='Present'
            ).count()
            absent_count = total_classes - present_count
            percentage = round((present_count / total_classes) * 100, 1) if total_classes > 0 else 0

            session_records = []
            for session in sessions:
                record = AttendanceRecord.objects.filter(session=session, student=student_profile).first()
                session_records.append({
                    'date': session.created_at,
                    'subject': subject.name,
                    'status': record.status if record else 'Absent',
                    'face_verified': record.face_verified if record else False,
                    'location_verified': record.location_verified if record else False,
                })

            semester_data['subjects'].append({
                'subject': subject,
                'total_classes': total_classes,
                'present_count': present_count,
                'absent_count': absent_count,
                'percentage': percentage,
                'session_records': session_records,
            })

        attendance_details.append(semester_data)

    export = request.GET.get('export')
    if export == 'excel':
        return export_student_attendance_excel(student_profile, course, attendance_details)
    elif export == 'pdf':
        return export_student_attendance_pdf(student_profile, course, attendance_details)

    context = {
        'student_profile': student_profile,
        'course': course,
        'attendance_details': attendance_details,
    }
    return render(request, 'attendance/student_attendance_detail.html', context)


def export_student_attendance_excel(student_profile, course, attendance_details):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Attendance"

    ws.merge_cells('A1:F1')
    ws['A1'] = f"Student: {student_profile.user.get_full_name()} ({student_profile.roll_no})"
    ws['A1'].font = Font(size=14, bold=True)
    ws['A2'] = f"Course: {course.code} - {course.name} | Department: {student_profile.department}"
    ws['A3'] = f"Email: {student_profile.user.email} | Phone: {student_profile.user.phone_number or 'N/A'}"
    ws['A4'] = ""

    row = 5
    for sem_data in attendance_details:
        ws.merge_cells(f'A{row}:F{row}')
        ws[f'A{row}'] = f"Semester {sem_data['semester']} - Section {sem_data['section'].name}"
        ws[f'A{row}'].font = Font(size=12, bold=True, color="FFFFFF")
        ws[f'A{row}'].fill = PatternFill(start_color="1a237e", end_color="1a237e", fill_type="solid")
        row += 1

        for sub_data in sem_data['subjects']:
            ws[f'A{row}'] = f"{sub_data['subject'].code} - {sub_data['subject'].name}"
            ws[f'B{row}'] = f"Total: {sub_data['total_classes']}"
            ws[f'C{row}'] = f"Present: {sub_data['present_count']}"
            ws[f'D{row}'] = f"Absent: {sub_data['absent_count']}"
            ws[f'E{row}'] = f"%: {sub_data['percentage']}%"
            ws[f'F{row}'] = ""
            for col in range(1, 7):
                cell = ws.cell(row=row, column=col)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="e0e0e0", end_color="e0e0e0", fill_type="solid")
            row += 1

            headers = ['Date', 'Status', 'Face', 'Location']
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
                cell.fill = PatternFill(start_color="d0d0d0", end_color="d0d0d0", fill_type="solid")
            row += 1

            for record in sub_data['session_records']:
                ws.cell(row=row, column=1, value=record['date'].strftime('%d %b %Y %I:%M %p') if record['date'] else '')
                ws.cell(row=row, column=2, value=record['status'])
                ws.cell(row=row, column=3, value='✓' if record['face_verified'] else '✗')
                ws.cell(row=row, column=4, value='✓' if record['location_verified'] else '✗')
                row += 1

            row += 1
        row += 1

    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except Exception:
                pass
        adjusted_width = min(max_length + 2, 30)
        ws.column_dimensions[col_letter].width = adjusted_width

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=student_attendance_{student_profile.roll_no}_{course.code}.xlsx'
    wb.save(response)
    return response


def export_student_attendance_pdf(student_profile, course, attendance_details):
    html = f"""
    <html>
    <head>
        <style>
            body {{ font-family: 'Helvetica', sans-serif; margin: 20px; }}
            h2 {{ color: #1a237e; }}
            .student-info {{ margin-bottom: 20px; }}
            .semester {{ margin-top: 25px; }}
            .subject {{ margin-top: 15px; border: 1px solid #ddd; padding: 10px; }}
            .subject h5 {{ margin: 0 0 5px 0; }}
            .summary {{ display: flex; gap: 15px; margin-bottom: 8px; }}
            .summary-item {{ background: #f5f5f5; padding: 4px 10px; border-radius: 4px; }}
            table {{ width: 100%; border-collapse: collapse; margin-top: 8px; }}
            th {{ background-color: #1a237e; color: white; padding: 5px; text-align: left; font-size: 10px; }}
            td {{ padding: 4px 6px; border: 1px solid #ddd; font-size: 9px; }}
            tr:nth-child(even) {{ background-color: #f9f9f9; }}
            .badge-success {{ background: #28a745; color: white; padding: 1px 8px; border-radius: 12px; }}
            .badge-danger {{ background: #dc3545; color: white; padding: 1px 8px; border-radius: 12px; }}
        </style>
    </head>
    <body>
        <div class="student-info">
            <h2>Student Attendance Detail</h2>
            <p><strong>Name:</strong> {student_profile.user.get_full_name()} ({student_profile.roll_no})</p>
            <p><strong>Course:</strong> {course.code} - {course.name} | <strong>Department:</strong> {student_profile.department}</p>
            <p><strong>Email:</strong> {student_profile.user.email} | <strong>Phone:</strong> {student_profile.user.phone_number or 'N/A'}</p>
        </div>
    """

    for sem_data in attendance_details:
        html += f"""
        <div class="semester">
            <h4>Semester {sem_data['semester']} - Section {sem_data['section'].name}</h4>
        """
        for sub_data in sem_data['subjects']:
            html += f"""
            <div class="subject">
                <h5>{sub_data['subject'].code} - {sub_data['subject'].name}</h5>
                <div class="summary">
                    <span class="summary-item">Total: {sub_data['total_classes']}</span>
                    <span class="summary-item">Present: {sub_data['present_count']}</span>
                    <span class="summary-item">Absent: {sub_data['absent_count']}</span>
                    <span class="summary-item">%: {sub_data['percentage']}%</span>
                </div>
                <table>
                    <thead><tr><th>Date</th><th>Status</th><th>Face</th><th>Location</th></tr></thead>
                    <tbody>
            """
            for record in sub_data['session_records']:
                status_class = 'badge-success' if record['status'] == 'Present' else 'badge-danger'
                face = '✓' if record['face_verified'] else '✗'
                loc = '✓' if record['location_verified'] else '✗'
                html += f"""
                        <tr>
                            <td>{record['date'].strftime('%d %b %Y %I:%M %p') if record['date'] else ''}</td>
                            <td><span class="{status_class}">{record['status']}</span></td>
                            <td>{face}</td>
                            <td>{loc}</td>
                        </tr>
                """
            html += """
                    </tbody>
                </table>
            </div>
            """
        html += "</div>"

    html += """
    </body>
    </html>
    """

    result = io.BytesIO()
    pdf = pisa.pisaDocument(io.BytesIO(html.encode("UTF-8")), result)
    if pdf.err:
        return HttpResponse('Error generating PDF', status=500)

    response = HttpResponse(result.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename=student_attendance_{student_profile.roll_no}_{course.code}.pdf'
    return response


# =============================================================================
# STAFF LIST & DETAIL (PRINCIPAL)
# =============================================================================

@login_required
def principal_teaching_staff(request):
    if request.user.user_type != 'principal':
        return redirect('home')

    staff = User.objects.filter(user_type__in=['lecturer', 'hod']).select_related('assigned_course')

    search = request.GET.get('search', '')
    course_code = request.GET.get('course', '')
    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')
    export = request.GET.get('export')

    if search:
        staff = staff.filter(
            Q(first_name__icontains=search) |
            Q(last_name__icontains=search) |
            Q(username__icontains=search) |
            Q(email__icontains=search)
        )
    if course_code:
        staff = staff.filter(assigned_course__code=course_code)

    today = timezone.now().date()
    if not from_date:
        from_date = today.replace(day=1).isoformat()
    if not to_date:
        to_date = today.isoformat()

    try:
        from_dt = datetime.strptime(from_date, '%Y-%m-%d').date()
        to_dt = datetime.strptime(to_date, '%Y-%m-%d').date()
    except ValueError:
        from_dt = today.replace(day=1)
        to_dt = today

    all_dates = [from_dt + timedelta(days=i) for i in range((to_dt - from_dt).days + 1)]
    holiday_dates = HolidaySetting.objects.filter(
        date__gte=from_dt,
        date__lte=to_dt,
        status='holiday'
    ).values_list('date', flat=True)
    holiday_dates = list(holiday_dates)
    working_dates = [d for d in all_dates if d not in holiday_dates]
    total_working_days = len(working_dates)

    staff_data = []
    for user in staff:
        present_count = StaffAttendance.objects.filter(
            user=user,
            date__in=working_dates,
            status='present'
        ).count()
        percentage = round((present_count / total_working_days) * 100, 1) if total_working_days > 0 else 0.0
        staff_data.append({
            'user': user,
            'percentage': percentage,
            'present_count': present_count,
            'total_working_days': total_working_days,
        })

    if export == 'excel':
        return export_staff_excel(staff_data, from_date, to_date)
    elif export == 'pdf':
        return export_staff_pdf(staff_data, from_date, to_date)

    courses = Course.objects.all().order_by('code')
    context = {
        'staff_data': staff_data,
        'type': 'Teaching',
        'courses': courses,
        'search': search,
        'selected_course': course_code,
        'from_date': from_date,
        'to_date': to_date,
        'total_working_days': total_working_days,
        'total_holidays': len(holiday_dates),
    }
    return render(request, 'attendance/principal_teaching_staff_list.html', context)


def export_staff_excel(staff_data, from_date, to_date):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Teaching Staff"

    headers = ['Name', 'Username', 'Role', 'Assigned Course', 'Attendance %', 'Present / Total']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.font = Font(color="FFFFFF", bold=True)

    for row, entry in enumerate(staff_data, start=2):
        ws.cell(row=row, column=1, value=entry['user'].get_full_name())
        ws.cell(row=row, column=2, value=entry['user'].username)
        ws.cell(row=row, column=3, value=entry['user'].get_user_type_display())
        ws.cell(row=row, column=4, value=entry['user'].assigned_course.code if entry['user'].assigned_course else '—')
        ws.cell(row=row, column=5, value=entry['percentage'])
        ws.cell(row=row, column=6, value=f"{entry['present_count']} / {entry['total_working_days']}")

    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except Exception:
                pass
        adjusted_width = min(max_length + 2, 30)
        ws.column_dimensions[col_letter].width = adjusted_width

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=teaching_staff_{from_date}_to_{to_date}.xlsx'
    wb.save(response)
    return response


def export_staff_pdf(staff_data, from_date, to_date):
    html = f"""
    <html>
    <head>
        <style>
            body {{ font-family: 'Helvetica', sans-serif; }}
            h2 {{ color: #1a2a6c; }}
            table {{ width: 100%; border-collapse: collapse; margin-top: 20px; }}
            th {{ background-color: #1a2a6c; color: white; padding: 8px; text-align: left; }}
            td {{ padding: 6px 8px; border: 1px solid #ddd; }}
            tr:nth-child(even) {{ background-color: #f8f9fa; }}
            .summary {{ margin-bottom: 15px; }}
        </style>
    </head>
    <body>
        <h2>Teaching Staff Attendance</h2>
        <div class="summary">
            <strong>Date Range:</strong> {from_date} to {to_date}<br>
            <strong>Total Working Days:</strong> {staff_data[0]['total_working_days'] if staff_data else 0}<br>
            <strong>Total Staff:</strong> {len(staff_data)}
        </div>
        <table>
            <thead>
                <tr>
                    <th>Name</th>
                    <th>Username</th>
                    <th>Role</th>
                    <th>Assigned Course</th>
                    <th>Attendance %</th>
                    <th>Present / Total</th>
                </tr>
            </thead>
            <tbody>
    """

    for entry in staff_data:
        html += f"""
                <tr>
                    <td>{entry['user'].get_full_name()}</td>
                    <td>{entry['user'].username}</td>
                    <td>{entry['user'].get_user_type_display()}</td>
                    <td>{entry['user'].assigned_course.code if entry['user'].assigned_course else '—'}</td>
                    <td>{entry['percentage']}%</td>
                    <td>{entry['present_count']} / {entry['total_working_days']}</td>
                </tr>
        """

    html += """
            </tbody>
        </table>
    </body>
    </html>
    """

    result = io.BytesIO()
    pdf = pisa.pisaDocument(io.BytesIO(html.encode("UTF-8")), result)
    if pdf.err:
        return HttpResponse('Error generating PDF', status=500)

    response = HttpResponse(result.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename=teaching_staff_{from_date}_to_{to_date}.pdf'
    return response


# accounts/views.py (or wherever principal_non_teaching_staff is defined)

@login_required
def principal_non_teaching_staff(request):
    if request.user.user_type != 'principal':
        return redirect('home')

    # Base queryset: only users with user_type='staff' (Non‑Teaching Staff)
    staff_qs = User.objects.filter(user_type='staff').select_related('assigned_course')

    # Filters
    search = request.GET.get('search', '')
    designation_filter = request.GET.get('designation', '')
    from_date = request.GET.get('from_date', '')
    to_date = request.GET.get('to_date', '')

    if search:
        staff_qs = staff_qs.filter(
            Q(first_name__icontains=search) |
            Q(last_name__icontains=search) |
            Q(username__icontains=search) |
            Q(email__icontains=search)
        )

    if designation_filter:
        user_ids = StaffProfile.objects.filter(designation=designation_filter).values_list('user_id', flat=True)
        staff_qs = staff_qs.filter(id__in=user_ids)

    staff_data = []
    for user in staff_qs:
        try:
            designation = user.staff_profile.designation or 'Not set'
        except StaffProfile.DoesNotExist:
            designation = 'Not set'

        attendances = StaffAttendance.objects.filter(user=user)
        if from_date:
            attendances = attendances.filter(date__gte=from_date)
        if to_date:
            attendances = attendances.filter(date__lte=to_date)

        total = attendances.count()
        present = attendances.filter(status='present').count()
        percentage = round((present / total) * 100) if total > 0 else 0

        staff_data.append({
            'user': user,
            'designation': designation,
            'percentage': percentage,
            'present_count': present,
            'total_working_days': total,
        })

    # --- Export handling ---
    export = request.GET.get('export')
    if export == 'excel':
        return export_non_teaching_staff_excel(staff_data, from_date, to_date)
    elif export == 'pdf':
        return export_non_teaching_staff_pdf(staff_data, from_date, to_date)

    # Distinct designations for filter dropdown
    designations = StaffProfile.objects.filter(
        user__in=staff_qs
    ).values_list('designation', flat=True).distinct().exclude(designation__isnull=True).exclude(designation='')

    context = {
        'staff_data': staff_data,
        'designations': designations,
        'selected_designation': designation_filter,
        'search': search,
        'from_date': from_date,
        'to_date': to_date,
        'courses': Course.objects.all().order_by('code'),  # if needed
    }
    return render(request, 'attendance/principal_non_teaching_staff_list.html', context)

import io
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from django.http import HttpResponse
from xhtml2pdf import pisa

def export_non_teaching_staff_excel(staff_data, from_date, to_date):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Non-Teaching Staff"

    # Headers
    headers = ['Name', 'Username', 'Designation', 'Attendance %', 'Present / Total']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal='center')
        cell.fill = PatternFill(start_color="1a2a6c", end_color="1a2a6c", fill_type="solid")

    # Data rows
    for row, entry in enumerate(staff_data, start=2):
        ws.cell(row=row, column=1, value=entry['user'].get_full_name())
        ws.cell(row=row, column=2, value=entry['user'].username)
        ws.cell(row=row, column=3, value=entry['designation'])
        ws.cell(row=row, column=4, value=entry['percentage'])
        ws.cell(row=row, column=5, value=f"{entry['present_count']} / {entry['total_working_days']}")

    # Auto-size columns
    for col in ws.columns:
        max_length = 0
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 30)
        ws.column_dimensions[col_letter].width = adjusted_width

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=non_teaching_staff_{from_date or "all"}_to_{to_date or "all"}.xlsx'
    wb.save(response)
    return response


def export_non_teaching_staff_pdf(staff_data, from_date, to_date):
    html = f"""
    <html>
    <head>
        <style>
            body {{ font-family: 'Helvetica', sans-serif; margin: 20px; }}
            h2 {{ color: #1a2a6c; }}
            .summary {{ margin-bottom: 15px; }}
            table {{ width: 100%; border-collapse: collapse; margin-top: 20px; }}
            th {{ background-color: #1a2a6c; color: white; padding: 8px; text-align: left; }}
            td {{ padding: 6px 8px; border: 1px solid #ddd; }}
            tr:nth-child(even) {{ background-color: #f8f9fa; }}
        </style>
    </head>
    <body>
        <h2>Non-Teaching Staff Attendance</h2>
        <div class="summary">
            <strong>Date Range:</strong> {from_date or 'Start'} to {to_date or 'End'}<br>
            <strong>Total Staff:</strong> {len(staff_data)}
        </div>
        <table>
            <thead>
                <tr>
                    <th>Name</th>
                    <th>Username</th>
                    <th>Designation</th>
                    <th>Attendance %</th>
                    <th>Present / Total</th>
                </tr>
            </thead>
            <tbody>
    """
    for entry in staff_data:
        html += f"""
                <tr>
                    <td>{entry['user'].get_full_name()}</td>
                    <td>{entry['user'].username}</td>
                    <td>{entry['designation']}</td>
                    <td>{entry['percentage']}%</td>
                    <td>{entry['present_count']} / {entry['total_working_days']}</td>
                </tr>
        """
    html += """
            </tbody>
        </table>
    </body>
    </html>
    """

    result = io.BytesIO()
    pdf = pisa.pisaDocument(io.BytesIO(html.encode("UTF-8")), result)
    if pdf.err:
        return HttpResponse('Error generating PDF', status=500)

    response = HttpResponse(result.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename=non_teaching_staff_{from_date or "all"}_to_{to_date or "all"}.pdf'
    return response
   
@login_required
def principal_staff_detail(request, user_id):
    if request.user.user_type != 'principal':
        return redirect('home')

    staff = get_object_or_404(User, id=user_id)

    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')

    if from_date and from_date.lower() == 'none':
        from_date = None
    if to_date and to_date.lower() == 'none':
        to_date = None

    attendance_qs = StaffAttendance.objects.filter(user=staff).order_by('-date')
    if from_date:
        try:
            from_dt = datetime.strptime(from_date, '%Y-%m-%d').date()
            attendance_qs = attendance_qs.filter(date__gte=from_dt)
        except ValueError:
            pass
    if to_date:
        try:
            to_dt = datetime.strptime(to_date, '%Y-%m-%d').date()
            attendance_qs = attendance_qs.filter(date__lte=to_dt)
        except ValueError:
            pass
    attendance_records = attendance_qs

    sessions = AttendanceSession.objects.filter(lecturer=staff).order_by('-created_at') if staff.user_type in ['lecturer', 'hod'] else []

    export = request.GET.get('export')
    if export == 'excel':
        return export_staff_attendance_excel(staff, attendance_records, from_date, to_date)
    elif export == 'pdf':
        return export_staff_attendance_pdf(staff, attendance_records, from_date, to_date)

    context = {
        'staff': staff,
        'attendance_records': attendance_records,
        'sessions': sessions,
        'from_date': from_date,
        'to_date': to_date,
    }
    return render(request, 'attendance/principal_teaching_staff_detail.html', context)


def export_staff_attendance_excel(staff, records, from_date, to_date):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Attendance"

    ws.merge_cells('A1:F1')
    ws['A1'] = f"Staff Attendance: {staff.get_full_name()} ({staff.username})"
    ws['A1'].font = Font(size=14, bold=True)
    ws['A2'] = f"Date Range: {from_date or 'Start'} to {to_date or 'End'}"
    ws['A3'] = f"Role: {staff.get_user_type_display()} | Email: {staff.email}"
    ws['A4'] = ""

    headers = ['Date', 'Time In', 'Time Out', 'Status', 'Face Verified', 'Location Verified']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=5, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal='center')
        cell.fill = PatternFill(start_color="1a237e", end_color="1a237e", fill_type="solid")

    row = 6
    for rec in records:
        ws.cell(row=row, column=1, value=rec.date.strftime('%d %b %Y') if rec.date else '')
        ws.cell(row=row, column=2, value=rec.time_in.strftime('%I:%M %p') if rec.time_in else '')
        ws.cell(row=row, column=3, value=rec.time_out.strftime('%I:%M %p') if rec.time_out else '—')
        ws.cell(row=row, column=4, value=rec.status.title())
        ws.cell(row=row, column=5, value='✓' if rec.face_verified else '✗')
        ws.cell(row=row, column=6, value='✓' if rec.location_verified else '✗')
        row += 1

    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except Exception:
                pass
        adjusted_width = min(max_length + 2, 30)
        ws.column_dimensions[col_letter].width = adjusted_width

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=staff_attendance_{staff.username}_{from_date or "all"}_{to_date or "all"}.xlsx'
    wb.save(response)
    return response


def export_staff_attendance_pdf(staff, records, from_date, to_date):
    html = f"""
    <html>
    <head>
        <style>
            body {{ font-family: 'Helvetica', sans-serif; margin: 20px; }}
            h2 {{ color: #1a237e; }}
            .info {{ margin-bottom: 20px; }}
            table {{ width: 100%; border-collapse: collapse; }}
            th {{ background-color: #1a237e; color: white; padding: 6px; text-align: left; font-size: 10px; }}
            td {{ padding: 4px 6px; border: 1px solid #ddd; font-size: 9px; }}
            tr:nth-child(even) {{ background-color: #f9f9f9; }}
        </style>
    </head>
    <body>
        <div class="info">
            <h2>Staff Attendance - {staff.get_full_name()}</h2>
            <p><strong>Username:</strong> {staff.username} | <strong>Role:</strong> {staff.get_user_type_display()}</p>
            <p><strong>Email:</strong> {staff.email} | <strong>Phone:</strong> {staff.phone_number or 'N/A'}</p>
            <p><strong>Date Range:</strong> {from_date or 'Start'} to {to_date or 'End'}</p>
        </div>
        <table>
            <thead>
                <tr>
                    <th>Date</th>
                    <th>Time In</th>
                    <th>Time Out</th>
                    <th>Status</th>
                    <th>Face</th>
                    <th>Location</th>
                </tr>
            </thead>
            <tbody>
    """

    for rec in records:
        face = '✓' if rec.face_verified else '✗'
        loc = '✓' if rec.location_verified else '✗'
        html += f"""
                <tr>
                    <td>{rec.date.strftime('%d %b %Y') if rec.date else ''}</td>
                    <td>{rec.time_in.strftime('%I:%M %p') if rec.time_in else ''}</td>
                    <td>{rec.time_out.strftime('%I:%M %p') if rec.time_out else '—'}</td>
                    <td>{rec.status.title()}</td>
                    <td>{face}</td>
                    <td>{loc}</td>
                </tr>
        """

    html += """
            </tbody>
        </table>
    </body>
    </html>
    """

    result = io.BytesIO()
    pdf = pisa.pisaDocument(io.BytesIO(html.encode("UTF-8")), result)
    if pdf.err:
        return HttpResponse('Error generating PDF', status=500)

    response = HttpResponse(result.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename=staff_attendance_{staff.username}_{from_date or "all"}_{to_date or "all"}.pdf'
    return response


@login_required
def principal_non_teaching_staff_detail(request, staff_id):
    if request.user.user_type != 'principal':
        return redirect('home')

    staff = get_object_or_404(User, id=staff_id, user_type='staff')

    try:
        staff_profile = StaffProfile.objects.get(user=staff)
        designation = staff_profile.designation or 'Not set'
        employee_id = staff_profile.employee_id or 'Not set'
        department = staff_profile.department or 'Not set'
    except StaffProfile.DoesNotExist:
        designation = 'Not set'
        employee_id = 'Not set'
        department = 'Not set'

    attendance_records = StaffAttendance.objects.filter(user=staff).order_by('-date', '-time_in')

    export_type = request.GET.get('export')
    if export_type == 'excel':
        return export_attendance_excel(staff, attendance_records, designation)
    elif export_type == 'pdf':
        return export_attendance_pdf(staff, attendance_records, designation)

    context = {
        'staff': staff,
        'designation': designation,
        'employee_id': employee_id,
        'department': department,
        'attendance_records': attendance_records,
    }
    return render(request, 'attendance/principal_non_teaching_staff_detail.html', context)


def export_attendance_excel(staff, records, designation):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Attendance'

    ws.merge_cells('A1:F1')
    title_cell = ws['A1']
    title_cell.value = f'Attendance Report - {staff.get_full_name()} ({designation})'
    title_cell.font = Font(size=14, bold=True)
    title_cell.alignment = Alignment(horizontal='center')

    headers = ['Date', 'Time In', 'Time Out', 'Status', 'Face Verified', 'Location Verified']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='1a2a6c', end_color='1a2a6c', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    row_num = 4
    for rec in records:
        ws.cell(row=row_num, column=1, value=rec.date.strftime('%d %b %Y'))
        ws.cell(row=row_num, column=2, value=rec.time_in.strftime('%H:%M') if rec.time_in else '')
        ws.cell(row=row_num, column=3, value=rec.time_out.strftime('%H:%M') if rec.time_out else '')
        ws.cell(row=row_num, column=4, value=rec.status.title())
        ws.cell(row=row_num, column=5, value='Yes' if rec.face_verified else 'No')
        ws.cell(row=row_num, column=6, value='Yes' if rec.location_verified else 'No')
        row_num += 1

    for col in range(1, 7):
        ws.column_dimensions[chr(64 + col)].width = 18

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="attendance_{staff.username}_{datetime.now().strftime("%Y%m%d")}.xlsx"'
    wb.save(response)
    return response


def export_attendance_pdf(staff, records, designation):
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="attendance_{staff.username}_{datetime.now().strftime("%Y%m%d")}.pdf"'

    doc = SimpleDocTemplate(
        response,
        pagesize=landscape(A4),
        rightMargin=30,
        leftMargin=30,
        topMargin=30,
        bottomMargin=30
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('TitleStyle', parent=styles['Heading1'], fontSize=16, alignment=1, spaceAfter=12)
    normal_style = styles['Normal']

    elements = []

    title = f"Attendance Report - {staff.get_full_name()} ({designation})"
    elements.append(Paragraph(title, title_style))
    elements.append(Spacer(1, 0.2 * inch))

    data = [['Date', 'Time In', 'Time Out', 'Status', 'Face', 'Location']]
    for rec in records:
        data.append([
            rec.date.strftime('%d %b %Y'),
            rec.time_in.strftime('%H:%M') if rec.time_in else '',
            rec.time_out.strftime('%H:%M') if rec.time_out else '',
            rec.status.title(),
            '✓' if rec.face_verified else '✗',
            '✓' if rec.location_verified else '✗',
        ])

    table = Table(data, colWidths=[1.0 * inch, 0.8 * inch, 0.8 * inch, 1.0 * inch, 0.8 * inch, 0.8 * inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a2a6c')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    elements.append(table)

    doc.build(elements)
    return response


@login_required
def principal_search_students(request):
    if request.user.user_type != 'principal':
        return redirect('home')

    students = StudentProfile.objects.select_related('user')

    course_id = request.GET.get('course')
    semester = request.GET.get('semester')
    section_id = request.GET.get('section')
    name = request.GET.get('name', '')
    roll_no = request.GET.get('roll_no', '')

    if course_id:
        students = students.filter(studentenrollment__course_id=course_id).distinct()
    if semester:
        students = students.filter(studentenrollment__semester=semester).distinct()
    if section_id:
        students = students.filter(studentenrollment__section_id=section_id).distinct()
    if name:
        students = students.filter(
            Q(user__first_name__icontains=name) |
            Q(user__last_name__icontains=name) |
            Q(user__username__icontains=name)
        )
    if roll_no:
        students = students.filter(roll_no__icontains=roll_no)

    courses = Course.objects.all().order_by('code')
    sections = Section.objects.all().order_by('name')
    semesters = range(1, 9)

    context = {
        'students': students,
        'courses': courses,
        'sections': sections,
        'semesters': semesters,
        'selected_course': course_id,
        'selected_semester': semester,
        'selected_section': section_id,
        'search_name': name,
        'search_roll': roll_no,
    }
    return render(request, 'attendance/principal_search_students.html', context)


# =============================================================================
# HOD VIEWS
# =============================================================================

@login_required
def hod_faculty_list(request):
    if request.user.user_type != 'hod':
        messages.error(request, 'Access denied.')
        return redirect('home')

    course = request.user.assigned_course
    if not course:
        messages.error(request, 'No course assigned to you.')
        return redirect('hod_dashboard')

    faculty = User.objects.filter(user_type='lecturer', assigned_course=course)
    search = request.GET.get('search', '')
    designation = request.GET.get('designation', '')

    if search:
        faculty = faculty.filter(
            Q(first_name__icontains=search) |
            Q(last_name__icontains=search) |
            Q(username__icontains=search) |
            Q(email__icontains=search)
        )
    if designation:
        faculty = faculty.filter(staff_profile__designation=designation)

    designations = StaffProfile.objects.filter(
        user__in=faculty
    ).values_list('designation', flat=True).distinct().exclude(designation='')

    context = {
        'faculty': faculty,
        'course': course,
        'search': search,
        'designations': designations,
        'selected_designation': designation,
    }
    return render(request, 'attendance/hod_faculty_list.html', context)


# =============================================================================
# SESSION MANAGEMENT
# =============================================================================

@login_required
def start_session(request):
    if request.user.user_type != 'lecturer':
        return redirect('student_dashboard')

    courses = Course.objects.all()

    if request.method == 'POST':
        course_id = request.POST.get('course')
        semester = request.POST.get('semester')
        section_value = request.POST.get('section')   # can be 'all' or an ID
        subject_id = request.POST.get('subject')
        latitude = request.POST.get('latitude')
        longitude = request.POST.get('longitude')
        radius_feet = request.POST.get('radius_feet')

        course = get_object_or_404(Course, id=course_id)
        subject = get_object_or_404(Subject, id=subject_id)

        # Handle section
        if section_value == 'all':
            section = None   # NULL means "All Sections"
        else:
            section = get_object_or_404(Section, id=section_value)

        session = AttendanceSession.objects.create(
            lecturer=request.user,
            course=course,
            semester=int(semester),
            section=section,
            subject=subject,
            latitude=float(latitude),
            longitude=float(longitude),
            radius_feet=float(radius_feet)
        )

        return redirect('session_monitor', session_id=session.id)

    context = {'courses': courses}
    return render(request, 'attendance/start_session.html', context)


def mark_absent_for_missing_students(self):
    """Create Absent records for students who haven't marked attendance."""
    # Get all student IDs enrolled in this course & semester
    enrolled_qs = StudentEnrollment.objects.filter(
        course=self.course,
        semester=self.semester,
        subjects=self.subject
    )
    if self.section:
        enrolled_qs = enrolled_qs.filter(section=self.section)

    enrolled_student_ids = enrolled_qs.values_list('student_id', flat=True)

    existing_student_ids = AttendanceRecord.objects.filter(
        session=self
    ).values_list('student_id', flat=True)

    missing_ids = set(enrolled_student_ids) - set(existing_student_ids)

    for student_id in missing_ids:
        AttendanceRecord.objects.create(
            session=self,
            student_id=student_id,
            student_latitude=0.0,
            student_longitude=0.0,
            location_verified=False,
            face_verified=False,
            status='Absent'
        )
        

@login_required
def session_monitor(request, session_id):
    session = get_object_or_404(
        AttendanceSession.objects.select_related('course', 'section', 'subject', 'lecturer'),
        id=session_id
    )

    # Access control...
    if request.user.user_type not in ['lecturer', 'principal', 'admin', 'hod']:
        if request.user.user_type == 'student':
            return redirect('take_attendance', session_id=session_id)
        else:
            messages.error(request, 'Access denied.')
            return redirect('home')

    if request.user.user_type == 'lecturer' and request.user != session.lecturer:
        messages.error(request, 'You can only view your own sessions.')
        return redirect('lecturer_dashboard')

    # Get attendance records
    attendance_records = AttendanceRecord.objects.filter(session=session).select_related('student__user')
    present_count = attendance_records.filter(status='Present').count()
    absent_count = attendance_records.filter(status='Absent').count()

    # Enrollment count: handle all sections
    enrolled_qs = StudentEnrollment.objects.filter(
        course=session.course,
        semester=session.semester,
        subjects=session.subject
    )
    if session.section is None:
        enrolled_count = StudentEnrollment.objects.filter(
            course=session.course,
            semester=session.semester,
            subjects=session.subject
        ).values('student').distinct().count()
    else:
        enrolled_count = StudentEnrollment.objects.filter(
            course=session.course,
            semester=session.semester,
            section=session.section,
            subjects=session.subject
        ).values('student').distinct().count()

    attendance_records = AttendanceRecord.objects.filter(session=session).select_related('student__user')
    present_count = attendance_records.filter(status='Present').count()
    absent_count = attendance_records.filter(status='Absent').count()
    percentage = round((present_count / enrolled_count) * 100) if enrolled_count > 0 else 0

    context = {
        'session': session,
        'attendance_records': attendance_records,
        'enrolled_count': enrolled_count,
        'present_count': present_count,
        'absent_count': absent_count,
        'percentage': percentage,
    }

    return render(request, 'attendance/session_monitor.html', context)


@login_required
def take_attendance(request, session_id):
    session = get_object_or_404(AttendanceSession, id=session_id, is_active=True)
    student_profile = get_object_or_404(StudentProfile, user=request.user)
    already_marked = AttendanceRecord.objects.filter(session=session, student=student_profile).exists()

    # Enrollment check: if session.section is None (All Sections), ignore section in filter
    if session.section is None:
        is_enrolled = StudentEnrollment.objects.filter(
            student=student_profile,
            course=session.course,
            semester=session.semester,
            subjects=session.subject
        ).exists()
    else:
        is_enrolled = StudentEnrollment.objects.filter(
            student=student_profile,
            course=session.course,
            semester=session.semester,
            section=session.section,
            subjects=session.subject
        ).exists()

    context = {
        'session': session,
        'already_marked': already_marked,
        'is_enrolled': is_enrolled,
        'student': request.user,
        'student_profile': student_profile,
    }
    return render(request, 'attendance/take_attendance.html', context)


@login_required
def end_session(request, session_id):
    session = get_object_or_404(AttendanceSession, id=session_id, lecturer=request.user)
    session.end_session()
    session.mark_absent_for_missing_students()
    return redirect('lecturer_dashboard')


@login_required
def get_session_records(request, session_id):
    session = get_object_or_404(AttendanceSession, id=session_id)

    if request.user != session.lecturer:
        return JsonResponse({'error': 'Unauthorized'}, status=403)

    records = AttendanceRecord.objects.filter(
        session=session
    ).select_related('student__user').values(
        'student__roll_no',
        'student__user__first_name',
        'student__user__last_name',
        'timestamp',
        'status',
        'face_verified',
        'location_verified'
    )

    return JsonResponse(list(records), safe=False)


# =============================================================================
# AJAX ENDPOINTS
# =============================================================================

@login_required
def get_sections(request, course_id, semester):
    sections = Section.objects.filter(course_id=course_id, semester=semester).values('id', 'name')
    return JsonResponse(list(sections), safe=False)


@login_required
def get_subjects(request, course_id, semester):
    subjects = Subject.objects.filter(course_id=course_id, semester=semester).values('id', 'name', 'code')
    return JsonResponse(list(subjects), safe=False)


# =============================================================================
# LECTURER & STAFF ATTENDANCE (SELF)
# =============================================================================

@login_required
def lecturer_mark_attendance(request):
    if request.user.user_type != 'lecturer':
        messages.error(request, 'Access denied.')
        return redirect('home')
    return render(request, 'attendance/lecturer_mark_attendance.html')


@login_required
def staff_dashboard(request):
    if request.user.user_type != 'staff':
        messages.error(request, 'Access denied.')
        return redirect('home')

    today = timezone.now().date()
    marked_today = StaffAttendance.objects.filter(user=request.user, date=today).exists()

    context = {'marked_today': marked_today}
    return render(request, 'attendance/staff_dashboard.html', context)


@login_required
def staff_mark_attendance(request):
    if request.user.user_type != 'staff':
        messages.error(request, 'Access denied.')
        return redirect('home')
    return render(request, 'attendance/staff_mark_attendance.html')


@login_required
def staff_attendance_history(request):
    if request.user.user_type != 'staff':
        messages.error(request, 'Access denied.')
        return redirect('home')

    records = StaffAttendance.objects.filter(
        user=request.user
    ).order_by('-date', '-time_in')

    context = {'records': records}
    return render(request, 'attendance/staff_attendance_history.html', context)


@login_required
def lecturer_history(request):
    if request.user.user_type != 'lecturer':
        messages.error(request, 'Access denied.')
        return redirect('home')
    return render(request, 'attendance/lecturer_history.html')


@login_required
def lecturer_my_attendance(request):
    if request.user.user_type != 'lecturer':
        messages.error(request, 'Access denied.')
        return redirect('home')

    records = StaffAttendance.objects.filter(user=request.user).order_by('-date', '-time_in')
    context = {
        'records': records,
        'type': 'My',
    }
    return render(request, 'attendance/lecturer_attendance_list.html', context)


from django.db.models import Count, Q, OuterRef, Subquery

@login_required
def lecturer_student_sessions(request):
    if request.user.user_type != 'lecturer':
        messages.error(request, 'Access denied.')
        return redirect('home')

    sessions = AttendanceSession.objects.filter(
        lecturer=request.user
    ).select_related('course', 'subject', 'section').order_by('-created_at')

    # For each session, compute totals manually
    session_data = []
    for session in sessions:
        # Get enrolled students count
        enrolled_qs = StudentEnrollment.objects.filter(
            course=session.course,
            semester=session.semester,
            subjects=session.subject
        )
        if session.section:
            enrolled_qs = enrolled_qs.filter(section=session.section)
        total_students = enrolled_qs.values('student').distinct().count()

        # Present and absent from AttendanceRecord
        records = AttendanceRecord.objects.filter(session=session)
        present_count = records.filter(status='Present').count()
        absent_count = records.filter(status='Absent').count()

        session_data.append({
            'session': session,
            'total_students': total_students,
            'present_count': present_count,
            'absent_count': absent_count,
        })

    context = {'sessions': session_data}
    return render(request, 'attendance/lecturer_student_sessions.html', context)

@login_required
def lecturer_session_detail(request, session_id):
    if request.user.user_type != 'lecturer':
        messages.error(request, 'Access denied.')
        return redirect('home')

    session = get_object_or_404(
        AttendanceSession.objects.select_related('course', 'subject', 'section'),
        id=session_id,
        lecturer=request.user
    )

    records = AttendanceRecord.objects.filter(
        session=session
    ).select_related('student__user').order_by('student__roll_no')

    # Calculate total enrolled
    if session.section is None:
        total_enrolled = StudentEnrollment.objects.filter(
            course=session.course,
            semester=session.semester,
            subjects=session.subject
        ).values('student').distinct().count()
    else:
        total_enrolled = StudentEnrollment.objects.filter(
            course=session.course,
            semester=session.semester,
            section=session.section,
            subjects=session.subject
        ).values('student').distinct().count()

    present = records.filter(status='Present').count()
    absent = records.filter(status='Absent').count()

    context = {
        'session': session,
        'records': records,
        'total': records.count(),
        'total_enrolled': total_enrolled,
        'present': present,
        'absent': absent,
    }
    return render(request, 'attendance/lecturer_session_detail.html', context)

# =============================================================================
# ADMIN: HOLIDAY CALENDAR
# =============================================================================

@login_required
def admin_holiday_calendar(request, year=None, month=None):
    if request.user.user_type != 'admin':
        messages.error(request, 'Access denied.')
        return redirect('admin_dashboard')

    if year is None:
        year = datetime.now().year
    if month is None:
        month = datetime.now().month

    if request.method == 'POST':
        for key, value in request.POST.items():
            if key.startswith('day_status_'):
                date_str = key.replace('day_status_', '')
                try:
                    date_obj = datetime.strptime(date_str, '%Y-%m-%d').date()
                except ValueError:
                    continue

                if value == 'holiday':
                    holiday, created = HolidaySetting.objects.get_or_create(date=date_obj)
                    holiday.status = 'holiday'
                    holiday.save()
                elif value == 'workday':
                    holiday, created = HolidaySetting.objects.get_or_create(date=date_obj)
                    holiday.status = 'working'
                    holiday.save()
                else:
                    HolidaySetting.objects.filter(date=date_obj).delete()

        messages.success(request, 'Holiday settings updated successfully.')
        return redirect('admin_dashboard')

    first_day = datetime(year, month, 1)
    if month == 12:
        last_day = datetime(year + 1, 1, 1) - timedelta(days=1)
    else:
        last_day = datetime(year, month + 1, 1) - timedelta(days=1)

    holidays = HolidaySetting.objects.filter(
        date__gte=first_day.date(),
        date__lte=last_day.date()
    )
    holiday_dict = {h.date.strftime('%Y-%m-%d'): h for h in holidays}

    total_days = (last_day - first_day).days + 1
    holiday_count = sum(1 for h in holiday_dict.values() if h.status == 'holiday')
    workday_count = sum(1 for h in holiday_dict.values() if h.status == 'working')
    unset_count = total_days - holiday_count - workday_count

    cal = calendar.monthcalendar(year, month)
    month_name = calendar.month_name[month]

    calendar_grid = []
    for week in cal:
        week_data = []
        for day in week:
            if day == 0:
                week_data.append({'empty': True})
            else:
                date_str = f"{year:04d}-{month:02d}-{day:02d}"
                holiday = holiday_dict.get(date_str)
                week_data.append({
                    'day': day,
                    'date_str': date_str,
                    'is_holiday': holiday.status == 'holiday' if holiday else False,
                    'empty': False,
                })
        calendar_grid.append(week_data)

    prev_month = month - 1 if month > 1 else 12
    prev_year = year if month > 1 else year - 1
    next_month = month + 1 if month < 12 else 1
    next_year = year if month < 12 else year + 1

    context = {
        'year': year,
        'month': month,
        'month_name': month_name,
        'calendar_grid': calendar_grid,
        'holiday_dict': holiday_dict,
        'total_days': total_days,
        'holiday_count': holiday_count,
        'workday_count': workday_count,
        'unset_count': unset_count,
        'prev_month': prev_month,
        'prev_year': prev_year,
        'next_month': next_month,
        'next_year': next_year,
    }
    return render(request, 'admin/admin_holiday_calendar.html', context)


@admin_required
def admin_toggle_holiday(request):
    if request.user.user_type != 'admin':
        return JsonResponse({'error': 'Unauthorized'}, status=403)

    if request.method == 'POST':
        date_str = request.POST.get('date')
        try:
            date_obj = datetime.strptime(date_str, '%Y-%m-%d').date()
        except ValueError:
            return JsonResponse({'error': 'Invalid date format'}, status=400)

        holiday, created = HolidaySetting.objects.get_or_create(date=date_obj)
        if created:
            holiday.status = 'holiday'
        else:
            holiday.status = 'working' if holiday.status == 'holiday' else 'holiday'
        holiday.save()

        return JsonResponse({
            'success': True,
            'date': date_str,
            'status': holiday.status,
            'is_holiday': holiday.status == 'holiday',
            'reason': getattr(holiday, 'reason', ''),
        })

    return JsonResponse({'error': 'Invalid method'}, status=405)


@admin_required
def admin_set_holiday_reason(request):
    if request.user.user_type != 'admin':
        return JsonResponse({'error': 'Unauthorized'}, status=403)

    if request.method == 'POST':
        date_str = request.POST.get('date')
        reason = request.POST.get('reason', '')
        try:
            date_obj = datetime.strptime(date_str, '%Y-%m-%d').date()
        except ValueError:
            return JsonResponse({'error': 'Invalid date format'}, status=400)

        holiday, created = HolidaySetting.objects.get_or_create(date=date_obj)
        holiday.reason = reason
        if created:
            holiday.status = 'holiday'
        holiday.save()

        return JsonResponse({'success': True, 'date': date_str, 'reason': reason})

    return JsonResponse({'error': 'Invalid method'}, status=405)

# Add this helper
def get_subjects_for_section(course, semester, section):
    """Return all subjects (core + languages) for a given section."""
    from accounts.models import Subject
    from django.db import models

    # Core subjects (based on section)
    core = get_mandatory_subjects(course, semester, section)

    # Language subjects (all languages for that course/semester)
    language_patterns = ['GEN', 'GHN', 'GKA', 'KAN', 'HIN']
    lang_subjects = Subject.objects.filter(course=course, semester=semester)
    query = models.Q()
    for p in language_patterns:
        query |= models.Q(code__icontains=p)
    lang_subjects = lang_subjects.filter(query)

    # Combine and remove duplicates
    all_subjects = list(core) + list(lang_subjects)
    seen = set()
    unique = []
    for s in all_subjects:
        if s.id not in seen:
            seen.add(s.id)
            unique.append(s)
    return unique


@login_required
def get_subjects(request, course_id, semester):
    course = get_object_or_404(Course, id=course_id)
    section_id = request.GET.get('section')

    if section_id and section_id != 'all':
        section = get_object_or_404(Section, id=section_id)
        subjects = get_subjects_for_section(course, semester, section)
    else:
        # 'all' or no section → all subjects for the course/semester
        subjects = Subject.objects.filter(course_id=course_id, semester=semester)

    data = [{'id': s.id, 'name': s.name, 'code': s.code} for s in subjects]
    return JsonResponse(data, safe=False)